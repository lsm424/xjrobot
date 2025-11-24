# from tools import list_all_tools, call_tool_by_name

# # 列出所有可用工具
# print("可用工具列表：")
# tools = list_all_tools()
# for tool in tools:
#     print(f"- {tool['name']}: {tool['description']}")
#     print(f"  参数: {[param['name'] for param in tool['parameters']]}")

# # 测试计算器工具
# print("\n测试计算器工具：")
# try:
#     result_add = call_tool_by_name("add", 5, 3)
#     print(f"5 + 3 = {result_add}")
    
#     result_subtract = call_tool_by_name("subtract", 10, 4)
#     print(f"10 - 4 = {result_subtract}")
    
#     result_multiply = call_tool_by_name("multiply", 6, 7)
#     print(f"6 * 7 = {result_multiply}")
    
#     result_divide = call_tool_by_name("divide", 20, 4)
#     print(f"20 / 4 = {result_divide}")
    
#     # 测试除零错误
#     # result_error = call_tool_by_name("divide", 10, 0)
# except ValueError as e:
#     print(f"错误: {e}")
import openpyxl

def process_schedule(file_path):
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # 非文化课列表
    non_cultural_courses = ['音', '体', '美', '信']
    
    # 获取班级列表（从第3列开始，第1行）
    class_names = []
    col = 3
    while ws.cell(row=1, column=col).value:
        class_names.append(ws.cell(row=1, column=col).value)
        col += 1
    
    results = []
    
    # 遍历每个班级（列）
    for class_idx, class_name in enumerate(class_names, start=3):
        # 获取周二第三节课内容（第4行，当前列）
        tuesday_third = ws.cell(row=4, column=class_idx).value
        
        # 如果是None，跳过
        if tuesday_third is None:
            continue
            
        # 判断是否为文化课（不在非文化课列表中）
        is_cultural = True
        for nc in non_cultural_courses:
            if nc in str(tuesday_third):
                is_cultural = False
                break
        
        if is_cultural:
            # 在周二其他节次寻找非文化课（第2-7行，除了第4行）
            found_tuesday = False
            replacement = None
            
            for row in range(2, 8):
                if row == 4:  # 跳过当前行（周二第三节）
                    continue
                    
                course = ws.cell(row=row, column=class_idx).value
                if course:
                    for nc in non_cultural_courses:
                        if nc in str(course):
                            period = row - 1  # 计算节次
                            replacement = f"周二第{period}节{course}"
                            found_tuesday = True
                            break
                    if found_tuesday:
                        break
            
            # 如果周二没找到，去周三找（第8-14行）
            if not found_tuesday:
                found_wednesday = False
                for row in range(8, 15):
                    course = ws.cell(row=row, column=class_idx).value
                    if course:
                        for nc in non_cultural_courses:
                            if nc in str(course):
                                period = row - 7  # 计算节次（周三第1节开始）
                                replacement = f"周三第{period}节{course}"
                                found_wednesday = True
                                break
                        if found_wednesday:
                            break
                
                if not found_wednesday:
                    # 周三也没找到，输出失败
                    results.append(f"{class_name}-周二第3节为{tuesday_third}-周二周三无法替换")
                else:
                    # 找到周三的替换
                    results.append(f"{class_name}-周二第3节为{tuesday_third}-换到-{replacement}")
            else:
                # 找到周二的替换
                results.append(f"{class_name}-周二第3节为{tuesday_third}-换到-{replacement}")
    
    return results

# 使用示例
if __name__ == "__main__":
    # 替换为你的文件路径
    file_path = "C:/Users/zhangyujie/Downloads/换课.xlsx"
    
    try:
        results = process_schedule(file_path)
        for result in results:
            print(result.replace("_x000D_",""))
    except FileNotFoundError:
        print(f"错误：找不到文件 '{file_path}'")
    except Exception as e:
        print(f"处理文件时出错：{e}")